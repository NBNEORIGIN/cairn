"""
Amazon Inventory Flatfile (.xlsm) parser.

Flatfile structure (confirmed from real data):
  - Sheet: "Template"
  - Row 1: Base64-encoded settings (ignore)
  - Row 2: Empty
  - Row 3: Reference group headers (ignore)
  - Row 4: Human-readable column headers ← use these
  - Row 5: API field names (ignore)
  - Row 6: Amazon example row (SKU = ABC123) ← skip
  - Row 7+: Real listing data

Column positions VARY between category templates. The parser uses
header name matching from Row 4, NOT column indices.

Multiple columns share the same header name:
  - 5x "Bullet Point"
  - 8x "Other Image URL"
  - 5x "Generic Keyword"
  - 5x "Recommended Browse Nodes"

We resolve these by ordinal (left-to-right occurrence count).
"""
import io
import json
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from core.amazon_intel.db import get_conn, insert_upload, update_upload


# Map our field names to (header_name, ordinal) tuples.
# Ordinal = 0-based index of occurrence for duplicate headers.
# Fields that have consistent header names across all templates
FIELD_MAP = {
    'title':            ('Title', 0),
    'sku':              ('SKU', 0),
    'product_id_type':  ('Product Id Type', 0),
    'product_id':       ('Product Id', 0),
    'item_name':        ('Item Name', 0),
    'brand':            ('Brand Name', 0),
    'parent_child':     ('Parentage Level', 0),
    'parent_sku':       ('Parent SKU', 0),
    'bullet1':          ('Bullet Point', 0),
    'bullet2':          ('Bullet Point', 1),
    'bullet3':          ('Bullet Point', 2),
    'bullet4':          ('Bullet Point', 3),
    'bullet5':          ('Bullet Point', 4),
    'description':      ('Product Description', 0),
    'keyword1':         ('Generic Keyword', 0),
    'keyword2':         ('Generic Keyword', 1),
    'keyword3':         ('Generic Keyword', 2),
    'keyword4':         ('Generic Keyword', 3),
    'keyword5':         ('Generic Keyword', 4),
    'main_image_url':   ('Main Image URL', 0),
    'image2':           ('Other Image URL', 0),
    'image3':           ('Other Image URL', 1),
    'image4':           ('Other Image URL', 2),
    'image5':           ('Other Image URL', 3),
    'image6':           ('Other Image URL', 4),
    'image7':           ('Other Image URL', 5),
    'image8':           ('Other Image URL', 6),
    'image9':           ('Other Image URL', 7),
    'swatch_image':     ('Swatch Image URL', 0),
    'colour':           ('Colour', 0),
    'size':             ('Size', 0),
    'material':         ('Material', 0),
    'browse_node_1':    ('Recommended Browse Nodes', 0),
    'browse_node_2':    ('Recommended Browse Nodes', 1),
}

# Fields where the header name includes marketplace-specific suffixes.
# These are resolved dynamically by _resolve_price_field().
PRICE_HEADER_PATTERNS = [
    'Your Price GBP',
    'Your Price USD',
    'Your Price',
]

FULFILMENT_HEADER_PATTERNS = [
    'Fulfillment Channel Code',
    'Fulfilment Channel Code',
]

RELEASE_DATE_HEADER_PATTERNS = [
    'Offering Release Date (Sell on Amazon',
    'Offering Release Date',
]


def _build_header_index(row4_values: list) -> dict[tuple[str, int], int]:
    """
    Build a mapping of (header_name, ordinal) -> column_index from Row 4.
    Handles duplicate header names by counting occurrences left to right.
    """
    counts: dict[str, int] = {}
    index: dict[tuple[str, int], int] = {}
    for col_idx, val in enumerate(row4_values):
        name = str(val).strip() if val else ''
        if not name:
            continue
        ordinal = counts.get(name, 0)
        counts[name] = ordinal + 1
        index[(name, ordinal)] = col_idx
    return index


def _resolve_field(field_name: str, row: list, header_index: dict) -> str | None:
    """Get a cell value by our field name, using the header index."""
    mapping = FIELD_MAP.get(field_name)
    if not mapping:
        return None
    col_idx = header_index.get(mapping)
    if col_idx is None:
        return None
    if col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None:
        return None
    return str(val).strip()


def _resolve_pattern_field(patterns: list[str], row: list,
                           header_index: dict) -> str | None:
    """
    Resolve a field where the header contains a variable suffix.
    E.g. 'Your Price GBP (Sell on Amazon, UK)' matches pattern 'Your Price GBP'.
    """
    for key, col_idx in header_index.items():
        header_name, ordinal = key
        if ordinal != 0:
            continue
        for pattern in patterns:
            if header_name.startswith(pattern):
                if col_idx < len(row) and row[col_idx] is not None:
                    return str(row[col_idx]).strip()
    return None


def _parse_price(val: str | None) -> float | None:
    if not val:
        return None
    val = val.replace(',', '').replace('£', '').replace('$', '').strip()
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def parse_flatfile(file_bytes: bytes, filename: str) -> list[dict]:
    """
    Parse an Amazon inventory flatfile (.xlsm/.xlsx).
    Returns list of parsed listing dicts.
    """
    # Write to temp file for openpyxl (it needs a file path for xlsm)
    suffix = '.xlsm' if filename.endswith('.xlsm') else '.xlsx'
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True, keep_links=False)

        if 'Template' not in wb.sheetnames:
            raise ValueError(f"No 'Template' sheet found. Sheets: {wb.sheetnames}")

        ws = wb['Template']
        rows_iter = ws.iter_rows(values_only=True)

        # Read rows 1-5 (skip them, but we need row 4 for headers)
        row1 = next(rows_iter)  # Base64 settings
        row2 = next(rows_iter)  # Empty
        row3 = next(rows_iter)  # Reference headers
        row4 = list(next(rows_iter))  # Human-readable headers �� use these
        row5 = next(rows_iter)  # API field names

        header_index = _build_header_index(row4)

        # Determine product_type from filename
        # e.g. "0_ANCHOR_STAKE-OFFICE_PRODUCTS.xlsm" -> "ANCHOR_STAKE-OFFICE_PRODUCTS"
        product_type = Path(filename).stem
        if product_type and product_type[0].isdigit() and '_' in product_type:
            product_type = product_type.split('_', 1)[1]

        listings = []
        skipped = 0

        for row_values in rows_iter:
            row = list(row_values)

            sku = _resolve_field('sku', row, header_index)
            if not sku:
                skipped += 1
                continue

            # Skip Amazon's example row (SKU = ABC123)
            if sku.upper() == 'ABC123':
                continue

            # Extract all fields
            asin = _resolve_field('product_id', row, header_index)
            product_id_type = _resolve_field('product_id_type', row, header_index)

            # Title: prefer Title, fall back to Item Name
            title = _resolve_field('title', row, header_index)
            if not title:
                title = _resolve_field('item_name', row, header_index)

            # Bullets
            bullets = []
            for i in range(1, 6):
                b = _resolve_field(f'bullet{i}', row, header_index)
                bullets.append(b)
            bullet_count = sum(1 for b in bullets if b)

            # Keywords
            keywords = []
            for i in range(1, 6):
                k = _resolve_field(f'keyword{i}', row, header_index)
                keywords.append(k)
            keyword_count = sum(1 for k in keywords if k)

            # Images
            main_image = _resolve_field('main_image_url', row, header_index)
            image_count = 1 if main_image else 0
            for i in range(2, 10):
                img = _resolve_field(f'image{i}', row, header_index)
                if img:
                    image_count += 1

            # Other fields
            price_raw = _resolve_pattern_field(PRICE_HEADER_PATTERNS, row, header_index)
            price = _parse_price(price_raw)
            parent_child = _resolve_field('parent_child', row, header_index)
            parent_sku = _resolve_field('parent_sku', row, header_index)
            brand = _resolve_field('brand', row, header_index)
            description = _resolve_field('description', row, header_index)
            colour = _resolve_field('colour', row, header_index)
            size = _resolve_field('size', row, header_index)
            material = _resolve_field('material', row, header_index)
            fulfilment_raw = _resolve_pattern_field(FULFILMENT_HEADER_PATTERNS, row, header_index)
            fulfilment = 'FBA' if fulfilment_raw and 'AMAZON' in fulfilment_raw.upper() else 'MFN'
            browse_1 = _resolve_field('browse_node_1', row, header_index)
            browse_2 = _resolve_field('browse_node_2', row, header_index)

            # Listing creation date — "Offering Release Date (Sell on Amazon, UK)"
            release_date_raw = _resolve_pattern_field(
                RELEASE_DATE_HEADER_PATTERNS, row, header_index
            )
            listing_created_at = None
            if release_date_raw:
                try:
                    from datetime import datetime
                    # Handle ISO 8601 format: 2023-08-29T14:44:37.267Z
                    listing_created_at = datetime.fromisoformat(
                        release_date_raw.replace('Z', '+00:00')
                    ).strftime('%Y-%m-%d %H:%M:%S')
                except (ValueError, AttributeError):
                    pass  # unparseable date — skip

            # Build raw_json with all non-empty cells for completeness
            raw = {}
            for col_idx, val in enumerate(row):
                if val is not None and str(val).strip():
                    header_name = row4[col_idx] if col_idx < len(row4) and row4[col_idx] else f'col_{col_idx}'
                    # Deduplicate by appending index
                    key = f'{header_name}_{col_idx}'
                    raw[key] = str(val).strip()

            listings.append({
                'sku': sku,
                'asin': asin if product_id_type and product_id_type.upper() == 'ASIN' else None,
                'product_id_type': product_id_type,
                'parent_child': parent_child,
                'parent_sku': parent_sku,
                'product_type': product_type,
                'title': title,
                'brand': brand,
                'bullet1': bullets[0], 'bullet2': bullets[1], 'bullet3': bullets[2],
                'bullet4': bullets[3], 'bullet5': bullets[4],
                'description': description,
                'keyword1': keywords[0], 'keyword2': keywords[1], 'keyword3': keywords[2],
                'keyword4': keywords[3], 'keyword5': keywords[4],
                'main_image_url': main_image,
                'image_count': image_count,
                'your_price': price,
                'fulfilment': fulfilment,
                'colour': colour,
                'size': size,
                'material': material,
                'browse_node_1': browse_1,
                'browse_node_2': browse_2,
                'keyword_count': keyword_count,
                'bullet_count': bullet_count,
                'listing_created_at': listing_created_at,
                'raw_json': raw,
            })

        wb.close()
        return listings

    finally:
        Path(tmp_path).unlink(missing_ok=True)


def parse_and_store_flatfile(file_bytes: bytes, filename: str,
                             marketplace: str = None) -> dict:
    """Parse a flatfile and store rows in ami_flatfile_data. Returns summary."""
    upload_id = insert_upload(filename, 'flatfile', marketplace)

    try:
        listings = parse_flatfile(file_bytes, filename)
    except Exception as e:
        update_upload(upload_id, error_count=1, errors=[str(e)], status='error')
        raise

    errors = []
    stored = 0

    with get_conn() as conn:
        with conn.cursor() as cur:
            for listing in listings:
                try:
                    cur.execute(
                        """INSERT INTO ami_flatfile_data
                               (upload_id, sku, asin, product_id_type, parent_child,
                                parent_sku, product_type, title, brand,
                                bullet1, bullet2, bullet3, bullet4, bullet5,
                                description, generic_keyword1, generic_keyword2,
                                generic_keyword3, generic_keyword4, generic_keyword5,
                                main_image_url, image_count, your_price, fulfilment,
                                colour, size, material, browse_node_1, browse_node_2,
                                keyword_count, bullet_count, listing_created_at, raw_json)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (upload_id, listing['sku'], listing['asin'],
                         listing['product_id_type'], listing['parent_child'],
                         listing['parent_sku'], listing['product_type'],
                         listing['title'], listing['brand'],
                         listing['bullet1'], listing['bullet2'], listing['bullet3'],
                         listing['bullet4'], listing['bullet5'],
                         listing['description'],
                         listing['keyword1'], listing['keyword2'],
                         listing['keyword3'], listing['keyword4'], listing['keyword5'],
                         listing['main_image_url'], listing['image_count'],
                         listing['your_price'], listing['fulfilment'],
                         listing['colour'], listing['size'], listing['material'],
                         listing['browse_node_1'], listing['browse_node_2'],
                         listing['keyword_count'], listing['bullet_count'],
                         listing['listing_created_at'],
                         json.dumps(listing['raw_json'])),
                    )
                    stored += 1
                except Exception as e:
                    errors.append(f"Row {listing['sku']}: {e}")

            conn.commit()

    # Also enrich SKU mapping from flatfile ASINs
    _enrich_sku_mapping(listings)

    update_upload(upload_id, row_count=stored, skip_count=len(listings) - stored,
                  error_count=len(errors), errors=errors[:50])

    return {
        'upload_id': upload_id,
        'filename': filename,
        'file_type': 'flatfile',
        'row_count': stored,
        'skip_count': len(listings) - stored,
        'error_count': len(errors),
        'errors': errors[:10],
        'status': 'complete',
    }


def _enrich_sku_mapping(listings: list[dict]):
    """
    Update ami_sku_mapping with ASIN data from flatfile.
    Flatfile ASINs are more current than the stock sheet for existing SKUs.
    For new SKUs not in the mapping, insert with source='flatfile'.
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            for listing in listings:
                if not listing['asin']:
                    continue
                # Try to update existing mapping with ASIN
                cur.execute(
                    """UPDATE ami_sku_mapping
                       SET asin = %s, updated_at = NOW()
                       WHERE sku = %s AND (asin IS NULL OR asin != %s)""",
                    (listing['asin'], listing['sku'], listing['asin']),
                )
                # If no row was updated (SKU not in mapping), insert if we can
                if cur.rowcount == 0:
                    # Check if already exists with same ASIN
                    cur.execute(
                        "SELECT id FROM ami_sku_mapping WHERE sku = %s",
                        (listing['sku'],),
                    )
                    if not cur.fetchone():
                        # New SKU — try to extract M-number from SKU
                        m_number = _extract_m_number(listing['sku'])
                        if m_number:
                            cur.execute(
                                """INSERT INTO ami_sku_mapping
                                       (sku, m_number, asin, source)
                                   VALUES (%s, %s, %s, 'flatfile')
                                   ON CONFLICT (sku) DO NOTHING""",
                                (listing['sku'], m_number, listing['asin']),
                            )
            conn.commit()


def _extract_m_number(sku: str) -> str | None:
    """
    Attempt to extract an M-number from a SKU string.
    Patterns: M0001, M0001US, M0001_F2CUS, etc.
    """
    import re
    match = re.match(r'^(M\d{4})', sku)
    if match:
        return match.group(1)
    return None
