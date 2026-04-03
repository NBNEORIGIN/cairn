"""
Amazon Advertising report parser.

Handles two report types:
  1. Search Term Report — per-keyword, needs ASIN mapping from ad group name
  2. Advertised Product Report — per-ASIN, direct mapping

Format: CSV or XLSX from Amazon Advertising Console.
"""
import csv
import io
import json
import tempfile
from pathlib import Path
from core.amazon_intel.db import get_conn, insert_upload, update_upload


# Header aliases for the Advertised Product Report
ADVERTISED_PRODUCT_ALIASES = {
    'asin': ['Advertised ASIN', 'ASIN', 'advertised_asin'],
    'sku': ['Advertised SKU', 'SKU', 'advertised_sku'],
    'campaign_name': ['Campaign Name', 'campaign_name'],
    'ad_group_name': ['Ad Group Name', 'ad_group_name'],
    'impressions': ['Impressions', 'impressions'],
    'clicks': ['Clicks', 'clicks'],
    'spend': ['Spend', 'Cost', 'spend'],
    'sales_7d': ['7 Day Total Sales', '7 Day Total Sales (£)',
                 'Sales', 'Total Sales', '7d Sales'],
    'orders_7d': ['7 Day Total Orders', 'Orders', '7d Orders',
                  '7 Day Total Orders (#)'],
    'acos': ['ACOS', 'Total Advertising Cost of Sales (ACOS)',
             'Advertising Cost of Sales'],
    'roas': ['ROAS', 'Total Return on Advertising Spend (ROAS)',
             'Return on Advertising Spend'],
}

# Header aliases for the Search Term Report
SEARCH_TERM_ALIASES = {
    'campaign_name': ['Campaign Name', 'campaign_name'],
    'ad_group_name': ['Ad Group Name', 'ad_group_name'],
    'targeting': ['Targeting', 'targeting', 'Keyword'],
    'match_type': ['Match Type', 'match_type'],
    'customer_search_term': ['Customer Search Term', 'customer_search_term',
                             'Search Term'],
    'impressions': ['Impressions', 'impressions'],
    'clicks': ['Clicks', 'clicks'],
    'spend': ['Spend', 'Cost', 'spend'],
    'sales_7d': ['7 Day Total Sales', 'Sales', '7d Sales',
                 '7 Day Total Sales (£)'],
    'orders_7d': ['7 Day Total Orders', 'Orders', '7d Orders'],
    'acos': ['ACOS', 'Total Advertising Cost of Sales (ACOS)'],
    'roas': ['ROAS', 'Total Return on Advertising Spend (ROAS)'],
}


def _build_column_map(headers: list[str], aliases: dict) -> dict[str, int]:
    col_map = {}
    for field, names in aliases.items():
        for alias in names:
            for i, h in enumerate(headers):
                if h.strip().lower() == alias.lower():
                    col_map[field] = i
                    break
            if field in col_map:
                break
    return col_map


def _clean_numeric(val: str) -> float | None:
    if not val or val.strip() in ('', '--', 'N/A'):
        return None
    val = val.strip().replace(',', '').replace('£', '').replace('$', '').replace('%', '')
    try:
        result = float(val)
        return result
    except (ValueError, TypeError):
        return None


def _clean_int(val: str) -> int:
    if not val or val.strip() in ('', '--', 'N/A'):
        return 0
    val = val.strip().replace(',', '')
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _detect_report_type(headers: list[str]) -> str:
    """Detect whether this is an advertised product or search term report."""
    header_lower = [h.strip().lower() for h in headers]
    if any('advertised asin' in h or 'advertised sku' in h for h in header_lower):
        return 'advertised_product'
    if any('search term' in h or 'customer search term' in h for h in header_lower):
        return 'search_term'
    if any('targeting' in h for h in header_lower):
        return 'search_term'
    # Default: if it has ASIN column, treat as advertised product
    if any('asin' in h for h in header_lower):
        return 'advertised_product'
    return 'search_term'


def _extract_asin_from_ad_group(ad_group_name: str) -> str | None:
    """
    Try to extract an ASIN from the ad group name.
    NBNE convention: ad group names often contain the ASIN or SKU.
    """
    import re
    # ASIN pattern: B followed by 9 alphanumeric chars
    # Can't use \b because underscores in campaign names are word chars
    match = re.search(r'(?:^|[^A-Z0-9])(B[0-9A-Z]{9})(?:[^A-Z0-9]|$)', ad_group_name or '')
    if match:
        return match.group(1)
    return None


def _read_file_content(content: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """Read CSV or XLSX content, return (headers, rows)."""
    if filename.endswith(('.xlsx', '.xlsm')):
        from openpyxl import load_workbook
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning)
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([str(v) if v is not None else '' for v in row])
            wb.close()
            if not all_rows:
                raise ValueError("Empty spreadsheet")
            return all_rows[0], all_rows[1:]
        finally:
            Path(tmp_path).unlink(missing_ok=True)
    else:
        # CSV
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"Cannot decode {filename}")

        reader = csv.reader(io.StringIO(text))
        headers = next(reader)
        rows = [line for line in reader if line and any(c.strip() for c in line)]
        return headers, rows


def parse_advertising_report(content: bytes, filename: str) -> list[dict]:
    """Parse an advertising report. Returns list of row dicts."""
    headers, data_rows = _read_file_content(content, filename)
    report_type = _detect_report_type(headers)

    aliases = ADVERTISED_PRODUCT_ALIASES if report_type == 'advertised_product' else SEARCH_TERM_ALIASES
    col_map = _build_column_map(headers, aliases)

    rows = []
    for line in data_rows:
        def _get(field: str) -> str:
            idx = col_map.get(field)
            if idx is None or idx >= len(line):
                return ''
            return line[idx].strip()

        asin = _get('asin') if report_type == 'advertised_product' else None
        ad_group = _get('ad_group_name')
        campaign = _get('campaign_name')

        # Extract ASIN from campaign or ad group name
        if not asin and campaign:
            asin = _extract_asin_from_ad_group(campaign)
        if not asin and ad_group:
            asin = _extract_asin_from_ad_group(ad_group)

        acos_raw = _clean_numeric(_get('acos'))
        if acos_raw and acos_raw > 1:
            acos_raw = acos_raw / 100.0

        roas_raw = _clean_numeric(_get('roas'))

        rows.append({
            'report_type': report_type,
            'campaign_name': _get('campaign_name') or None,
            'ad_group_name': ad_group or None,
            'asin': asin,
            'sku': _get('sku') if 'sku' in col_map else None,
            'targeting': _get('targeting') if 'targeting' in col_map else None,
            'match_type': _get('match_type') if 'match_type' in col_map else None,
            'customer_search_term': _get('customer_search_term') if 'customer_search_term' in col_map else None,
            'impressions': _clean_int(_get('impressions')),
            'clicks': _clean_int(_get('clicks')),
            'spend': _clean_numeric(_get('spend')),
            'sales_7d': _clean_numeric(_get('sales_7d')),
            'orders_7d': _clean_int(_get('orders_7d')),
            'acos': acos_raw,
            'roas': roas_raw,
        })

    return rows


def parse_and_store_advertising(content: bytes, filename: str,
                                 marketplace: str = None) -> dict:
    """Parse and store an advertising report. Returns summary."""
    upload_id = insert_upload(filename, 'advertising', marketplace)

    try:
        rows = parse_advertising_report(content, filename)
    except Exception as e:
        update_upload(upload_id, error_count=1, errors=[str(e)], status='error')
        raise

    errors = []
    stored = 0

    with get_conn() as conn:
        with conn.cursor() as cur:
            for row in rows:
                try:
                    cur.execute(
                        """INSERT INTO ami_advertising_data
                               (upload_id, report_type, campaign_name,
                                ad_group_name, asin, sku, targeting,
                                match_type, customer_search_term,
                                impressions, clicks, spend, sales_7d,
                                orders_7d, acos, roas)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (upload_id, row['report_type'], row['campaign_name'],
                         row['ad_group_name'], row['asin'], row['sku'],
                         row['targeting'], row['match_type'],
                         row['customer_search_term'],
                         row['impressions'], row['clicks'], row['spend'],
                         row['sales_7d'], row['orders_7d'],
                         row['acos'], row['roas']),
                    )
                    stored += 1
                except Exception as e:
                    errors.append(f"Row: {e}")

            conn.commit()

    update_upload(upload_id, row_count=stored, skip_count=len(rows) - stored,
                  error_count=len(errors), errors=errors[:50])

    return {
        'upload_id': upload_id,
        'filename': filename,
        'file_type': 'advertising',
        'row_count': stored,
        'skip_count': len(rows) - stored,
        'error_count': len(errors),
        'errors': errors[:10],
        'status': 'complete',
    }
